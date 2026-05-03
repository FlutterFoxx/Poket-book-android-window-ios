from dotenv import load_dotenv
load_dotenv()
import os, io, logging, secrets, string, base64
from fastapi import APIRouter, Request, HTTPException, Depends, Response
from typing import Optional, List
from bson import ObjectId
from datetime import datetime, timezone, timedelta
import uuid as uuid_lib
from core import *

router = APIRouter(prefix="/api")

# ─── Party Routes ─────────────────────────────────────────────────────────────

@router.get("/parties")
async def get_parties(current_user: dict = Depends(get_current_user)):
    uid = current_user["_id"]
    parties = await db.parties.find(
        {"user_id": uid, "is_deleted": {"$ne": True}}
    ).to_list(None)

    if not parties:
        return []

    party_ids = [str(p["_id"]) for p in parties]

    # ── Single aggregation: latest balance per party (replaces N get_party_balance calls) ──
    balance_pipeline = [
        {"$match": {"party_id": {"$in": party_ids}, "is_deleted": {"$ne": True}}},
        {"$sort": {"date": -1, "created_at": -1}},
        {"$group": {"_id": "$party_id", "balance": {"$first": "$balance"}}},
    ]
    balance_cursor = await db.ledger_entries.aggregate(balance_pipeline).to_list(None)
    balance_map = {row["_id"]: row["balance"] for row in balance_cursor}

    # ── Single aggregation: unlocked entry count per party (replaces N count_documents calls) ──
    unlocked_pipeline = [
        {"$match": {
            "party_id": {"$in": party_ids},
            "is_locked": {"$ne": True},
            "is_deleted": {"$ne": True},
        }},
        {"$group": {"_id": "$party_id", "count": {"$sum": 1}}},
    ]
    unlocked_cursor = await db.ledger_entries.aggregate(unlocked_pipeline).to_list(None)
    unlocked_map = {row["_id"]: row["count"] for row in unlocked_cursor}

    result = []
    for p in parties:
        pid = str(p["_id"])
        bal = balance_map.get(pid, 0.0)
        unlocked = unlocked_map.get(pid, 0)
        balance_zero = abs(round(bal, 2)) <= 0.01
        result.append({
            "id": pid, "name": p["name"],
            "mobile": p.get("mobile", ""), "address": p.get("address", ""),
            "current_balance": bal, "created_at": str(p.get("created_at", "")),
            "has_unlocked_entries": unlocked > 0,
            "is_deletable": balance_zero and unlocked == 0,
            "delete_reason": (
                "" if (balance_zero and unlocked == 0)
                else ("Balance zero nahi hai" if not balance_zero else f"{unlocked} unlocked entries hain")
            ),
        })
    return result

@router.post("/parties")
async def create_party(data: PartyCreate, current_user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    result = await db.parties.insert_one({
        "name": data.name, "mobile": data.mobile, "address": data.address,
        "user_id": current_user["_id"], "is_deleted": False,
        "created_at": now, "updated_at": now,
    })
    return {"id": str(result.inserted_id), "name": data.name, "mobile": data.mobile, "address": data.address, "current_balance": 0.0}

@router.get("/parties/{party_id}")
async def get_party(party_id: str, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"], "is_deleted": {"$ne": True}})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    bal = await get_party_balance(party_id)
    return {"id": party_id, "name": p["name"], "mobile": p.get("mobile", ""), "address": p.get("address", ""), "current_balance": bal}

@router.put("/parties/{party_id}")
async def update_party(party_id: str, data: PartyUpdate, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"], "is_deleted": {"$ne": True}})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    update = {k: v for k, v in data.model_dump().items() if v is not None}
    update["updated_at"] = datetime.now(timezone.utc)
    await db.parties.update_one({"_id": ObjectId(party_id)}, {"$set": update})
    updated = await db.parties.find_one({"_id": ObjectId(party_id)})
    bal = await get_party_balance(party_id)
    return {"id": party_id, "name": updated["name"], "mobile": updated.get("mobile", ""), "address": updated.get("address", ""), "current_balance": bal}

@router.delete("/parties/{party_id}")
async def delete_party(party_id: str, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    # Rule 1: Balance must be zero (within 1 paisa tolerance)
    balance = await get_party_balance(party_id)
    if abs(round(balance, 2)) > 0.01:
        bal_text = f"₹{abs(balance):.2f} {'Dena Hai' if balance > 0 else 'Lena Hai'}"
        raise HTTPException(status_code=400, detail=f"Balance clear karein pehle. Current: {bal_text}")
    # Rule 2: No unlocked entries
    unlocked = await db.ledger_entries.count_documents(
        {"party_id": party_id, "is_locked": {"$ne": True}, "is_deleted": {"$ne": True}}
    )
    if unlocked > 0:
        raise HTTPException(status_code=400, detail=f"{unlocked} unlocked entries hain. Pehle Tally/Lock karein.")
    # Soft delete
    await db.parties.update_one({"_id": ObjectId(party_id)}, {"$set": {"is_deleted": True, "updated_at": datetime.now(timezone.utc)}})
    return {"message": "Party deleted successfully"}

# ─── Ledger Routes ────────────────────────────────────────────────────────────

@router.get("/ledger/{party_id}/entries")
async def get_ledger_entries(party_id: str, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    entries = await db.ledger_entries.find(
        {"party_id": party_id, "is_deleted": {"$ne": True}},
        sort=[("date", 1), ("created_at", 1)]    # ASC — oldest first (top), newest at bottom
    ).to_list(None)
    # Batch-load counterparty names
    cp_ids = list({e["counterparty_id"] for e in entries if e.get("counterparty_id")})
    cp_map = {}
    if cp_ids:
        cps = await db.parties.find({"_id": {"$in": [ObjectId(cid) for cid in cp_ids]}}).to_list(None)
        cp_map = {str(c["_id"]): c["name"] for c in cps}
    formatted = []
    for e in entries:
        fe = format_entry(e)
        fe["counterparty_name"] = cp_map.get(e.get("counterparty_id", ""), "")
        formatted.append(fe)
    # current_balance = most recent entry (last in ASC order)
    current_balance = formatted[-1]["balance"] if formatted else 0.0
    return {
        "party": {"id": party_id, "name": p["name"], "mobile": p.get("mobile", ""), "address": p.get("address", "")},
        "entries": formatted,
        "current_balance": current_balance,
    }

@router.post("/ledger/{party_id}/entries")
async def add_ledger_entry(party_id: str, data: LedgerEntryCreate, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    # Validate counterparty if provided
    cp = None
    if data.counterparty_id and data.counterparty_id != party_id:
        cp = await db.parties.find_one({"_id": ObjectId(data.counterparty_id), "user_id": current_user["_id"], "is_deleted": {"$ne": True}})
        if not cp:
            raise HTTPException(status_code=404, detail="Counterparty not found")
    now = datetime.now(timezone.utc)
    transaction_id = str(uuid_lib.uuid4())
    # Entry A — Header party's ledger
    result_a = await db.ledger_entries.insert_one({
        "party_id": party_id, "counterparty_id": data.counterparty_id or None,
        "transaction_id": transaction_id, "date": data.date,
        "jama": data.jama, "naam": data.naam, "narration": data.narration,
        "balance": 0.0, "is_locked": False, "tally_id": None,
        "is_deleted": False, "created_at": now, "updated_at": now,
    })
    await recalculate_balances(party_id)
    # Entry B — Counterparty's ledger (mirrored: naam↔jama swapped)
    if cp:
        await db.ledger_entries.insert_one({
            "party_id": data.counterparty_id, "counterparty_id": party_id,
            "transaction_id": transaction_id, "date": data.date,
            "jama": data.naam,   # A's naam becomes B's jama
            "naam": data.jama,   # A's jama becomes B's naam
            "narration": data.narration,
            "balance": 0.0, "is_locked": False, "tally_id": None,
            "is_deleted": False, "created_at": now, "updated_at": now,
        })
        await recalculate_balances(data.counterparty_id)
    updated_a = await db.ledger_entries.find_one({"_id": result_a.inserted_id})
    fe = format_entry(updated_a)
    fe["counterparty_name"] = cp["name"] if cp else ""
    return fe

@router.put("/ledger/{party_id}/entries/{entry_id}")
async def update_ledger_entry(party_id: str, entry_id: str, data: LedgerEntryUpdate, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    e = await db.ledger_entries.find_one({"_id": ObjectId(entry_id), "party_id": party_id})
    if not e:
        raise HTTPException(status_code=404, detail="Entry not found")
    if e.get("is_locked"):
        raise HTTPException(status_code=403, detail="Locked entry edit nahi ho sakti")
    update_fields = {k: v for k, v in data.model_dump().items() if v is not None}
    update_fields["updated_at"] = datetime.now(timezone.utc)
    await db.ledger_entries.update_one({"_id": ObjectId(entry_id)}, {"$set": update_fields})
    await recalculate_balances(party_id)
    # Update mirrored entry (swap naam/jama)
    txn_id = e.get("transaction_id")
    if txn_id:
        mirror = await db.ledger_entries.find_one(
            {"transaction_id": txn_id, "party_id": {"$ne": party_id}, "is_deleted": {"$ne": True}}
        )
        if mirror and not mirror.get("is_locked"):
            mirror_update = {"updated_at": datetime.now(timezone.utc)}
            if "naam" in update_fields:
                mirror_update["jama"] = update_fields["naam"]
            if "jama" in update_fields:
                mirror_update["naam"] = update_fields["jama"]
            if "date" in update_fields:
                mirror_update["date"] = update_fields["date"]
            if "narration" in update_fields:
                mirror_update["narration"] = update_fields["narration"]
            await db.ledger_entries.update_one({"_id": mirror["_id"]}, {"$set": mirror_update})
            await recalculate_balances(mirror["party_id"])
    updated = await db.ledger_entries.find_one({"_id": ObjectId(entry_id)})
    fe = format_entry(updated)
    if e.get("counterparty_id"):
        cp = await db.parties.find_one({"_id": ObjectId(e["counterparty_id"])})
        fe["counterparty_name"] = cp["name"] if cp else ""
    return fe

@router.delete("/ledger/{party_id}/entries/{entry_id}")
async def delete_ledger_entry(party_id: str, entry_id: str, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    e = await db.ledger_entries.find_one({"_id": ObjectId(entry_id), "party_id": party_id})
    if not e:
        raise HTTPException(status_code=404, detail="Entry not found")
    if e.get("is_locked"):
        raise HTTPException(status_code=403, detail="Locked entry delete nahi ho sakti")
    now = datetime.now(timezone.utc)
    await db.ledger_entries.update_one({"_id": ObjectId(entry_id)}, {"$set": {"is_deleted": True, "updated_at": now}})
    await recalculate_balances(party_id)
    # Also delete the mirrored entry (same transaction_id)
    txn_id = e.get("transaction_id")
    if txn_id:
        mirror = await db.ledger_entries.find_one(
            {"transaction_id": txn_id, "party_id": {"$ne": party_id}, "is_deleted": {"$ne": True}}
        )
        if mirror and not mirror.get("is_locked"):
            await db.ledger_entries.update_one({"_id": mirror["_id"]}, {"$set": {"is_deleted": True, "updated_at": now}})
            await recalculate_balances(mirror["party_id"])
    return {"message": "Entry deleted successfully (both sides)"}

@router.post("/ledger/{party_id}/tally")
async def lock_party_entries(party_id: str, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    unlocked = await db.ledger_entries.find(
        {"party_id": party_id, "is_locked": {"$ne": True}, "is_deleted": {"$ne": True}}
    ).to_list(None)
    if not unlocked:
        raise HTTPException(status_code=400, detail="Koi unlocked entry nahi hai")
    closing_balance = await get_party_balance(party_id)
    now = datetime.now(timezone.utc)
    tally_result = await db.tallies.insert_one({
        "party_id": party_id, "user_id": current_user["_id"],
        "closing_balance": closing_balance, "timestamp": now, "type": "party",
    })
    tally_id = str(tally_result.inserted_id)
    # Lock this party's entries
    await db.ledger_entries.update_many(
        {"party_id": party_id, "is_locked": {"$ne": True}, "is_deleted": {"$ne": True}},
        {"$set": {"is_locked": True, "tally_id": tally_id}},
    )
    # Lock mirrored entries (same transaction_ids) in counterparties
    txn_ids = [e["transaction_id"] for e in unlocked if e.get("transaction_id")]
    if txn_ids:
        await db.ledger_entries.update_many(
            {"transaction_id": {"$in": txn_ids}, "party_id": {"$ne": party_id},
             "is_locked": {"$ne": True}, "is_deleted": {"$ne": True}},
            {"$set": {"is_locked": True, "tally_id": tally_id}},
        )
    return {"tally_id": tally_id, "closing_balance": closing_balance, "locked_count": len(unlocked), "timestamp": now.isoformat()}

@router.get("/ledger/{party_id}/tallies")
async def get_party_tallies(party_id: str, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    tallies = await db.tallies.find({"party_id": party_id}, sort=[("timestamp", -1)]).to_list(None)
    return [{"id": str(t["_id"]), "closing_balance": t.get("closing_balance", 0), "timestamp": str(t.get("timestamp", ""))} for t in tallies]

# ─── Balance Sheet ────────────────────────────────────────────────────────────

@router.get("/balance-sheet")
async def get_balance_sheet(current_user: dict = Depends(get_current_user)):
    parties = await db.parties.find({"user_id": current_user["_id"], "is_deleted": {"$ne": True}}).to_list(None)
    lena_hai, dena_hai = [], []
    for p in parties:
        pid = str(p["_id"])
        bal = await get_party_balance(pid)
        entry = {"id": pid, "name": p["name"], "mobile": p.get("mobile", ""), "balance": bal}
        # NAAM (credit) = DENA for party → positive balance → dena_hai (LEFT, BLUE)
        # JAMA (debit)  = LENA for party → negative balance → lena_hai (RIGHT, RED)
        if bal > 0:
            dena_hai.append({**entry, "amount": abs(bal)})   # positive = DENA (gave credit, will pay)
        elif bal < 0:
            lena_hai.append({**entry, "amount": abs(bal)})   # negative = LENA (received credit, will receive)
    total_dena = sum(x["amount"] for x in dena_hai)   # total of DENA parties
    total_lena = sum(x["amount"] for x in lena_hai)   # total of LENA parties
    return {
        "lena_hai": sorted(lena_hai, key=lambda x: x["amount"], reverse=True),
        "dena_hai": sorted(dena_hai, key=lambda x: x["amount"], reverse=True),
        "total_receivable": round(total_dena, 2),   # we will receive from DENA parties
        "total_payable": round(total_lena, 2),      # we will pay to LENA parties
        # net: negative = net LENA (we receive more) → formatBalance(negative) = Lena Hai RED ✓
        "net_balance": round(total_lena - total_dena, 2),
    }

# ─── Export Helper Functions ──────────────────────────────────────────────────

def fmt_inr(amount: float) -> str:
    return f"Rs.{abs(float(amount)):,.2f}"

def _format_balance_text(bal: float) -> str:
    """Format a balance value as human-readable Lena/Dena text."""
    if bal > 0:
        return f"{fmt_inr(bal)} Lena Hai"
    if bal < 0:
        return f"{fmt_inr(bal)} Dena Hai"
    return "Settled"

def _build_date_query(start_date: Optional[str], end_date: Optional[str]) -> dict:
    """Build a MongoDB date range sub-query."""
    if not start_date and not end_date:
        return {}
    date_q: dict = {}
    if start_date:
        date_q["$gte"] = start_date
    if end_date:
        date_q["$lte"] = end_date
    return {"date": date_q}

async def _load_entries_with_cp_names(party_id: str, extra_query: dict) -> tuple[list, dict]:
    """Fetch ledger entries and return (entries, counterparty_name_map)."""
    query = {"party_id": party_id, "is_deleted": {"$ne": True}, **extra_query}
    entries = await db.ledger_entries.find(
        query, sort=[("date", 1), ("created_at", 1)]
    ).to_list(None)
    cp_ids = list({e["counterparty_id"] for e in entries if e.get("counterparty_id")})
    cp_map: dict = {}
    if cp_ids:
        cps = await db.parties.find({"_id": {"$in": [ObjectId(c) for c in cp_ids]}}).to_list(None)
        cp_map = {str(c["_id"]): c["name"] for c in cps}
    return entries, cp_map

def _build_ledger_pdf_table_data(entries: list, cp_map: dict) -> list:
    """Convert ledger entries to a list of rows for the PDF table."""
    headers = ["Date", "Party Name", "Credit (नाम)", "Debit (जमा)", "Narration", "Balance", "Status"]
    rows = [headers]
    for e in entries:
        rows.append([
            e.get("date", ""),
            cp_map.get(e.get("counterparty_id", ""), "—"),
            fmt_inr(e.get("naam", 0)) if e.get("naam", 0) > 0 else "-",
            fmt_inr(e.get("jama", 0)) if e.get("jama", 0) > 0 else "-",
            str(e.get("narration", "") or "")[:35],
            _format_balance_text(e.get("balance", 0)),
            "Locked" if e.get("is_locked") else "Open",
        ])
    return rows

def _build_ledger_pdf_table_style():
    """Return the TableStyle for the ledger PDF table."""
    from reportlab.platypus import TableStyle
    from reportlab.lib import colors
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8B4513")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFE8CC"), colors.HexColor("#FFDAB0")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D6D3D1")),
        ("ALIGN", (2, 1), (3, -1), "RIGHT"),
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])

def _build_excel_ledger(party_name: str, date_range: str, entries: list, cp_map: dict):
    """Create and populate an openpyxl workbook for a ledger export."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(party_name)[:31]
    ws.append(["Statement", party_name])
    ws.append(["Period", date_range])
    ws.append(["Generated", datetime.now().strftime("%d-%m-%Y %H:%M")])
    ws.append([])
    headers = ["Date", "Party Name", "Credit/Naam", "Debit/Jama", "Narration", "Balance", "Balance Type", "Status"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8B4513")
    for e in entries:
        bal = e.get("balance", 0)
        bal_type = "Lena Hai" if bal > 0 else ("Dena Hai" if bal < 0 else "Settled")
        cp_name = cp_map.get(e.get("counterparty_id", ""), "")
        ws.append([
            e.get("date", ""), cp_name,
            e.get("naam", 0) or "", e.get("jama", 0) or "",
            e.get("narration", "") or "", abs(bal), bal_type,
            "Locked" if e.get("is_locked") else "Open",
        ])
    return wb

# ─── Export Routes (simplified using helpers above) ───────────────────────────

