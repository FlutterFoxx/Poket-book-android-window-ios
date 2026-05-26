from dotenv import load_dotenv
load_dotenv()
import os, io, logging, secrets, string, base64
from fastapi import APIRouter, Request, HTTPException, Depends, Response
from typing import Optional, List
from bson import ObjectId
from datetime import datetime, timezone, timedelta
import uuid as uuid_lib
from core import (db, get_current_user, get_current_user_dl, _fmt_dt, format_entry, get_party_balance, recalculate_balances)

router = APIRouter(prefix="/api")

# ─── Standalone balance sheet helper (for PDF/Excel export) ──────────────────
async def _fetch_balance_sheet(current_user: dict) -> dict:
    """Standalone version of get_balance_sheet for use in export endpoints."""
    parties = await db.parties.find({"user_id": current_user["_id"], "is_deleted": {"$ne": True}}).to_list(None)
    lena_hai, dena_hai = [], []
    for p in parties:
        pid = str(p["_id"])
        bal = await get_party_balance(pid)
        entry = {"id": pid, "name": p["name"], "mobile": p.get("mobile", ""), "balance": bal}
        if bal > 0:
            dena_hai.append({**entry, "amount": abs(bal)})
        elif bal < 0:
            lena_hai.append({**entry, "amount": abs(bal)})
    total_dena = sum(x["amount"] for x in dena_hai)
    total_lena = sum(x["amount"] for x in lena_hai)
    return {
        "lena_hai": sorted(lena_hai, key=lambda x: x["amount"], reverse=True),
        "dena_hai": sorted(dena_hai, key=lambda x: x["amount"], reverse=True),
        "total_receivable": round(total_dena, 2),
        "total_payable": round(total_lena, 2),
        "net_balance": round(total_lena - total_dena, 2),
    }


def fmt_inr(amount: float) -> str:
    return f"Rs.{abs(float(amount)):,.2f}"

def _format_balance_text(bal: float) -> str:
    """Format balance for PDF — just the amount, no Dena/Lena label."""
    if bal == 0:
        return "Settled"
    return fmt_inr(abs(bal))   # just the number, clean for sharing

def _build_date_query(start_date: Optional[str], end_date: Optional[str]) -> dict:
    """Build a MongoDB date range sub-query."""
    if not start_date and not end_date:
        return {}
    date_q: dict = {}
    if start_date:
        date_q["$gte"] = start_date
    if end_date:
        date_q["$lte"] = end_date
    return {"date": date_q}

async def _load_entries_with_cp_names(party_id: str, extra_query: dict) -> tuple[list, dict]:
    """Fetch ledger entries and return (entries, counterparty_name_map)."""
    query = {"party_id": party_id, "is_deleted": {"$ne": True}, **extra_query}
    entries = await db.ledger_entries.find(
        query, sort=[("date", 1), ("created_at", 1)]
    ).to_list(None)
    cp_ids = list({e["counterparty_id"] for e in entries if e.get("counterparty_id")})
    cp_map: dict = {}
    if cp_ids:
        cps = await db.parties.find({"_id": {"$in": [ObjectId(c) for c in cp_ids]}}).to_list(None)
        cp_map = {str(c["_id"]): c["name"] for c in cps}
    return entries, cp_map

def _build_ledger_pdf_table_data(entries: list, cp_map: dict) -> list:
    """Convert ledger entries to a list of rows for the PDF table."""
    headers = ["Date", "Party Name", "Credit (नाम)", "Debit (जमा)", "Narration", "Balance", "Status"]
    rows = [headers]
    for e in entries:
        rows.append([
            e.get("date", ""),
            cp_map.get(e.get("counterparty_id", ""), "—"),
            fmt_inr(e.get("naam", 0)) if e.get("naam", 0) > 0 else "-",
            fmt_inr(e.get("jama", 0)) if e.get("jama", 0) > 0 else "-",
            str(e.get("narration", "") or "")[:35],
            _format_balance_text(e.get("balance", 0)),
            "Locked" if e.get("is_locked") else "Open",
        ])
    return rows

def _build_ledger_pdf_table_style():
    """Return the TableStyle for the ledger PDF table."""
    from reportlab.platypus import TableStyle
    from reportlab.lib import colors
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8B4513")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFE8CC"), colors.HexColor("#FFDAB0")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D6D3D1")),
        ("ALIGN", (2, 1), (3, -1), "RIGHT"),
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])

def _build_excel_ledger(party_name: str, date_range: str, entries: list, cp_map: dict):
    """Create and populate an openpyxl workbook for a ledger export."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(party_name)[:31]
    ws.append(["Statement", party_name])
    ws.append(["Period", date_range])
    ws.append(["Generated", datetime.now().strftime("%d-%m-%Y %H:%M")])
    ws.append([])
    headers = ["Date", "Party Name", "Credit/Naam", "Debit/Jama", "Narration", "Balance", "Balance Type", "Status"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8B4513")
    for e in entries:
        bal = e.get("balance", 0)
        bal_type = "Settled" if bal == 0 else f"Rs.{abs(bal):,.2f}"
        cp_name = cp_map.get(e.get("counterparty_id", ""), "")
        ws.append([
            e.get("date", ""), cp_name,
            e.get("naam", 0) or "", e.get("jama", 0) or "",
            e.get("narration", "") or "", abs(bal), bal_type,
            "Locked" if e.get("is_locked") else "Open",
        ])
    return wb

# ─── Export Routes (simplified using helpers above) ───────────────────────────

@router.get("/export/ledger/{party_id}/pdf")
async def export_ledger_pdf(
    party_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user_dl)
):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_RIGHT, TA_LEFT, TA_CENTER
    from reportlab.lib.units import mm

    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")

    entries, cp_map = await _load_entries_with_cp_names(party_id, _build_date_query(start_date, end_date))
    date_range = f"{start_date or 'All entries'} to {end_date or 'Today'}"

    NAVY  = colors.HexColor("#0A1628");  GREEN = colors.HexColor("#22C55E")
    ORANGE = colors.HexColor("#FFB347"); GREY  = colors.HexColor("#9CA3AF")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=15*mm, rightMargin=15*mm, topMargin=10*mm, bottomMargin=12*mm)

    # Logo: full quality for high-resolution PDF (750KB+)
    LOGO_PATH = "/app/frontend/public/logo.png"
    try:
        from PIL import Image as PILImage
        pil_img = PILImage.open(LOGO_PATH).convert("RGBA")
        logo_buf = io.BytesIO()
        pil_img.save(logo_buf, format="PNG")  # full resolution, no resize
        logo_buf.seek(0)
        logo_img = RLImage(logo_buf, width=55, height=55)
    except Exception:
        logo_img = Paragraph("", ParagraphStyle("x"))

    hdr_tbl = Table([[
        logo_img,
        [Paragraph("<b>PoketBook</b>", ParagraphStyle("h1", fontSize=13, fontName="Helvetica-Bold", textColor=GREEN)),
         Paragraph("Digital Udhar Khaata", ParagraphStyle("h2", fontSize=8, textColor=colors.HexColor("#94A3B8")))],
        [Paragraph(f"Party Statement: <b>{p['name'].title()}</b>", ParagraphStyle("ht", fontSize=14, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_CENTER)),
         Paragraph(f"Period: {date_range}", ParagraphStyle("hp", fontSize=8, textColor=colors.HexColor("#CBD5E1"), alignment=TA_CENTER))],
        [Paragraph("<b>Flutter Fox</b>", ParagraphStyle("ff", fontSize=9, fontName="Helvetica-Bold", textColor=ORANGE, alignment=TA_RIGHT)),
         Paragraph("flutterfox.in", ParagraphStyle("ffu", fontSize=7, textColor=GREY, alignment=TA_RIGHT))],
    ]], colWidths=[18*mm, 45*mm, 130*mm, 40*mm])
    hdr_tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),NAVY), ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),8), ("BOTTOMPADDING",(0,0),(-1,-1),8),
        ("LEFTPADDING",(0,0),(0,0),8), ("RIGHTPADDING",(3,0),(3,0),8),
    ]))

    table_data = _build_ledger_pdf_table_data(entries, cp_map)
    t = Table(table_data, colWidths=[60, 80, 80, 80, 160, 110, 50])
    t.setStyle(_build_ledger_pdf_table_style())

    foot_style = ParagraphStyle("foot", fontSize=7, textColor=GREY, alignment=TA_CENTER)
    elements = [
        hdr_tbl, Spacer(1, 8), t, Spacer(1, 8),
        HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#E5E7EB")),
        Spacer(1, 3),
        Paragraph(f"PoketBook — Digital Udhar Khaata | Generated: {datetime.now().strftime('%d %b %Y %H:%M')} IST | poketbook.in", foot_style),
        Paragraph("Powered by Flutter Fox · flutterfox.in · New Delhi, India", foot_style),
    ]
    doc.build(elements)
    buf.seek(0)
    return Response(content=buf.read(), media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=PoketBook_{p['name'].title()}_Statement.pdf"})

@router.get("/export/ledger/{party_id}/excel")
async def export_ledger_excel(
    party_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user_dl)
):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")

    entries, cp_map = await _load_entries_with_cp_names(party_id, _build_date_query(start_date, end_date))
    date_range = f"{start_date or 'All'} to {end_date or 'Today'}"
    wb = _build_excel_ledger(p["name"], date_range, entries, cp_map)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=Statement_{p['name']}.xlsx"})
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    query = {"party_id": party_id, "is_deleted": {"$ne": True}}
    if start_date or end_date:
        date_q = {}
        if start_date:
            date_q["$gte"] = start_date
        if end_date:
            date_q["$lte"] = end_date
        query["date"] = date_q
    entries = await db.ledger_entries.find(query, sort=[("date", 1), ("created_at", 1)]).to_list(None)
    # Batch-load counterparty names
    cp_ids = list({e["counterparty_id"] for e in entries if e.get("counterparty_id")})
    cp_map = {}
    if cp_ids:
        cps = await db.parties.find({"_id": {"$in": [ObjectId(cid) for cid in cp_ids]}}).to_list(None)
        cp_map = {str(c["_id"]): c["name"] for c in cps}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(p["name"])[:31]
    ws.append(["Statement", p["name"]])
    ws.append(["Period", f"{start_date or 'All'} to {end_date or 'Today'}"])
    ws.append(["Generated", datetime.now().strftime("%d-%m-%Y %H:%M")])
    ws.append([])
    headers = ["Date", "Party Name", "Credit/Naam (लेना)", "Debit/Jama (देना)", "Narration", "Balance", "Balance Type", "Status"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8B4513")
    for e in entries:
        bal = e.get("balance", 0)
        bal_type = "Settled" if bal == 0 else f"Rs.{abs(bal):,.2f}"
        cp_name = cp_map.get(e.get("counterparty_id", ""), "")
        ws.append([e.get("date", ""), cp_name,
                   e.get("naam", 0) or "", e.get("jama", 0) or "",
                   e.get("narration", "") or "", abs(bal), bal_type,
                   "Locked" if e.get("is_locked") else "Open"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=Statement_{p['name']}.xlsx"})

@router.get("/export/balance-sheet/pdf")
async def export_bs_pdf(current_user: dict = Depends(get_current_user_dl)):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    data = await _fetch_balance_sheet(current_user)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=10*mm, bottomMargin=12*mm)

    BLUE   = colors.HexColor("#1E40AF");  BLUE_L = colors.HexColor("#DBEAFE");  BLUE_T = colors.HexColor("#1D4ED8")
    RED    = colors.HexColor("#991B1B");  RED_L  = colors.HexColor("#FEE2E2");  RED_T  = colors.HexColor("#B91C1C")
    NAVY   = colors.HexColor("#0A1628");  GREEN  = colors.HexColor("#22C55E")
    GREY   = colors.HexColor("#F9FAFB");  BORDER = colors.HexColor("#E5E7EB")

    dena = data["dena_hai"];  lena = data["lena_hai"];  max_rows = max(len(dena), len(lena), 1)
    net  = data.get("net_balance", 0)

    # ── Header table: full quality logo for 750KB+ PDF ──────────────────
    LOGO_PATH = "/app/frontend/public/logo.png"
    try:
        from PIL import Image as PILImage
        pil_img = PILImage.open(LOGO_PATH).convert("RGBA")
        lbuf = io.BytesIO(); pil_img.save(lbuf, format="PNG"); lbuf.seek(0)
        logo_cell = RLImage(lbuf, width=55, height=55)
    except Exception:
        logo_cell = Paragraph("", ParagraphStyle("x"))
    hdr_title  = Paragraph("<b>PoketBook</b>", ParagraphStyle("ht", fontSize=14, fontName="Helvetica-Bold", textColor=GREEN))
    hdr_sub    = Paragraph("Digital Udhar Khaata", ParagraphStyle("hs", fontSize=8, textColor=colors.HexColor("#94A3B8")))
    hdr_sheet  = Paragraph("Balance Sheet", ParagraphStyle("hsh", fontSize=16, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_CENTER))
    hdr_date   = Paragraph(datetime.now(timezone.utc).strftime("%d %B %Y"), ParagraphStyle("hd", fontSize=9, textColor=colors.HexColor("#CBD5E1"), alignment=TA_CENTER))
    ff_brand   = Paragraph("Powered by <b>Flutter Fox</b>", ParagraphStyle("ff", fontSize=8, textColor=colors.HexColor("#FFB347"), alignment=TA_RIGHT))
    ff_url     = Paragraph("flutterfox.in", ParagraphStyle("ffu", fontSize=7, textColor=colors.HexColor("#94A3B8"), alignment=TA_RIGHT))

    hdr_table = Table([
        [[logo_cell, hdr_title, hdr_sub],  [hdr_sheet, hdr_date],  [ff_brand, ff_url]]
    ], colWidths=[50*mm, 80*mm, 50*mm])
    hdr_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), NAVY),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (1,0), (1,0), "CENTER"), ("ALIGN", (2,0), (2,0), "RIGHT"),
        ("TOPPADDING", (0,0), (-1,-1), 8), ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING", (0,0), (0,0), 8), ("RIGHTPADDING", (2,0), (2,0), 8),
        ("INNERGRID", (0,0), (-1,-1), 0, NAVY),
    ]))

    # ── Summary strip ────────────────────────────────────────────────────────
    net_color = "#1E40AF" if net > 0 else ("#991B1B" if net < 0 else "#374151")
    net_label = "Net Payable" if net > 0 else ("Net Receivable" if net < 0 else "Balanced")
    summary = Table([[
        Paragraph(f"<b>Dena Hai (Payable):</b> ₹{fmt_inr(data['total_payable'])} [{len(dena)} parties]",
                  ParagraphStyle("s1", fontSize=8, textColor=BLUE_T)),
        Paragraph(f"<b>Lena Hai (Receivable):</b> ₹{fmt_inr(data['total_receivable'])} [{len(lena)} parties]",
                  ParagraphStyle("s2", fontSize=8, textColor=RED_T, alignment=TA_CENTER)),
        Paragraph(f"<b>{net_label}:</b> ₹{fmt_inr(abs(net))}",
                  ParagraphStyle("s3", fontSize=8, fontName="Helvetica-Bold",
                                 textColor=colors.HexColor(net_color), alignment=TA_RIGHT)),
    ]], colWidths=[60*mm, 70*mm, 50*mm])
    summary.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#F8FAFC")),
        ("BOX", (0,0), (-1,-1), 0.5, BORDER),
        ("TOPPADDING", (0,0), (-1,-1), 5), ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING", (0,0), (-1,-1), 8), ("RIGHTPADDING", (0,0), (-1,-1), 8),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))

    # ── Two-column ledger table ───────────────────────────────────────────────
    COL = [72*mm, 28*mm, 4*mm, 72*mm, 28*mm]
    hdr_s = ParagraphStyle("hdr", fontSize=9, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_LEFT)
    hdr_r = ParagraphStyle("hdr_r", fontSize=9, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_RIGHT)

    def _row(dp, da, lp, la, is_total=False):
        fn = "Helvetica-Bold" if is_total else "Helvetica"
        return [
            Paragraph(f'<font name="{fn}" color="{BLUE_T.hexval()}">{dp}</font>', ParagraphStyle("c",  fontSize=8, fontName=fn)),
            Paragraph(f'<font name="{fn}" color="{BLUE_T.hexval()}">{da}</font>', ParagraphStyle("r",  fontSize=8, fontName=fn, alignment=TA_RIGHT)),
            "", 
            Paragraph(f'<font name="{fn}" color="{RED_T.hexval()}">{lp}</font>',  ParagraphStyle("c2", fontSize=8, fontName=fn)),
            Paragraph(f'<font name="{fn}" color="{RED_T.hexval()}">{la}</font>',  ParagraphStyle("r2", fontSize=8, fontName=fn, alignment=TA_RIGHT)),
        ]

    tdata = [[Paragraph(f"DENA HAI / देना है  [{len(dena)} parties]", hdr_s), Paragraph("Amount (₹)", hdr_r),
              "", Paragraph(f"LENA HAI / लेना है  [{len(lena)} parties]", hdr_s), Paragraph("Amount (₹)", hdr_r)]]

    for i in range(max_rows):
        dp = dena[i] if i < len(dena) else {}; lp = lena[i] if i < len(lena) else {}
        tdata.append(_row(dp.get("name",""), fmt_inr(dp["amount"]) if dp.get("name") else "",
                          lp.get("name",""), fmt_inr(lp["amount"]) if lp.get("name") else ""))
    tdata.append(_row("TOTAL PAYABLE", fmt_inr(data["total_payable"]),
                      "TOTAL RECEIVABLE", fmt_inr(data["total_receivable"]), is_total=True))

    t = Table(tdata, colWidths=COL, repeatRows=1)
    row_styles = [s for i in range(1, max_rows+1) for s in [
        ("BACKGROUND",(0,i),(1,i), colors.white if i%2==1 else GREY),
        ("BACKGROUND",(3,i),(4,i), colors.white if i%2==1 else GREY),
    ]]
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(1,0),BLUE), ("BACKGROUND",(3,0),(4,0),RED),
        ("BACKGROUND",(2,0),(2,-1),colors.white),
        ("BOX",(0,0),(1,-1),0.5,BORDER), ("BOX",(3,0),(4,-1),0.5,BORDER),
        ("INNERGRID",(0,0),(1,-1),0.3,BORDER), ("INNERGRID",(3,0),(4,-1),0.3,BORDER),
        ("LINEABOVE",(0,-1),(1,-1),1.5,BLUE), ("LINEABOVE",(3,-1),(4,-1),1.5,RED),
        ("BACKGROUND",(0,-1),(1,-1),BLUE_L), ("BACKGROUND",(3,-1),(4,-1),RED_L),
        ("TOPPADDING",(0,0),(-1,-1),5), ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING",(0,0),(-1,-1),6), ("RIGHTPADDING",(0,0),(-1,-1),6),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"), *row_styles,
    ]))

    # ── Footer ────────────────────────────────────────────────────────────────
    footer_style = ParagraphStyle("foot", fontSize=7, textColor=colors.HexColor("#9CA3AF"), alignment=TA_CENTER)

    elements = [
        hdr_table, Spacer(1, 6), summary, Spacer(1, 8), t, Spacer(1, 8),
        HRFlowable(width="100%", thickness=0.5, color=BORDER),
        Spacer(1, 4),
        Paragraph("PoketBook — Digital Udhar Khaata for Indian Businesses | poketbook.in", footer_style),
        Paragraph("Powered by Flutter Fox · flutterfox.in · New Delhi, India", footer_style),
    ]
    doc.build(elements)
    buf.seek(0)
    return Response(content=buf.read(), media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=PoketBook_BalanceSheet.pdf"})

@router.get("/export/balance-sheet/excel")
async def export_bs_excel(current_user: dict = Depends(get_current_user_dl)):
    import openpyxl
    from openpyxl.styles import Font
    data = await _fetch_balance_sheet(current_user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.append(["Balance Sheet"])
    ws.append(["Generated", datetime.now().strftime("%d-%m-%Y %H:%M")])
    ws.append([])
    ws.append(["Lena Hai (Receivable)", "Amount", "", "Dena Hai (Payable)", "Amount"])
    for cell in ws[4]:
        cell.font = Font(bold=True)
    max_rows = max(len(data["lena_hai"]), len(data["dena_hai"]), 1)
    for i in range(max_rows):
        lena_row = data["lena_hai"][i] if i < len(data["lena_hai"]) else {}
        dena_row = data["dena_hai"][i] if i < len(data["dena_hai"]) else {}
        ws.append([lena_row.get("name", ""), lena_row.get("amount", "") or "", "", dena_row.get("name", ""), dena_row.get("amount", "") or ""])
    ws.append(["TOTAL RECEIVABLE", data["total_receivable"], "", "TOTAL PAYABLE", data["total_payable"]])
    ws.append(["NET BALANCE", data["net_balance"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=balance_sheet.xlsx"})


# ─── CSV/Google Sheets Backup ─────────────────────────────────────────────────

@router.get("/export/csv-backup")
async def export_csv_backup(current_user: dict = Depends(get_current_user)):
    """Export all user data as CSV — ready to import into Google Sheets"""
    import csv
    parties = await db.parties.find({"user_id": current_user["_id"], "is_deleted": {"$ne": True}}).to_list(None)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["=== PARTIES ==="])
    writer.writerow(["ID", "Name", "Mobile", "Address", "Current Balance"])
    for p in parties:
        pid = str(p["_id"])
        bal = await get_party_balance(pid)
        writer.writerow([pid, p["name"], p.get("mobile",""), p.get("address",""), bal])
    writer.writerow([])
    writer.writerow(["=== LEDGER ENTRIES ==="])
    writer.writerow(["Date", "Party", "Counterparty", "Credit/Naam", "Debit/Jama", "Narration", "Balance", "Locked"])
    for p in parties:
        pid = str(p["_id"])
        entries = await db.ledger_entries.find({"party_id": pid, "is_deleted": {"$ne": True}}, sort=[("date",1)]).to_list(None)
        cp_ids = list({e["counterparty_id"] for e in entries if e.get("counterparty_id")})
        cp_map = {}
        if cp_ids:
            cps = await db.parties.find({"_id": {"$in": [ObjectId(c) for c in cp_ids]}}).to_list(None)
            cp_map = {str(c["_id"]): c["name"] for c in cps}
        for e in entries:
            writer.writerow([e.get("date",""), p["name"], cp_map.get(e.get("counterparty_id",""),""),
                             e.get("naam",0), e.get("jama",0), e.get("narration",""),
                             e.get("balance",0), "Yes" if e.get("is_locked") else "No"])
    csv_content = buf.getvalue().encode("utf-8-sig")  # UTF-8 BOM for Excel/Sheets compatibility
    return Response(content=csv_content, media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=poketbook_backup_{datetime.now().strftime('%Y%m%d')}.csv"})

