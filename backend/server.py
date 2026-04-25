from dotenv import load_dotenv
load_dotenv()

import os
import io
import logging
import bcrypt
import jwt
import secrets
import string
import base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from fastapi import FastAPI, APIRouter, Request, HTTPException, Depends, Response
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel
from typing import Optional, List
from bson import ObjectId
from datetime import datetime, timezone, timedelta

import uuid as uuid_lib

mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

app = FastAPI()
api_router = APIRouter(prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=False,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory OTP store (replace with Redis in production)
_otp_store: dict = {}  # {phone: {otp, expires_at, attempts}}

GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/gmail.send",
    "openid",
    "https://www.googleapis.com/auth/userinfo.email",
]

SUBSCRIPTION_DAYS = {"trial": 7, "trial_ext": 15, "weekly": 7, "monthly": 30, "yearly": 365}
NEXT_BACKUP_DELTA = {"daily": timedelta(days=1), "weekly": timedelta(weeks=1), "monthly": timedelta(days=30)}
SUBSCRIPTION_PRICES = {"trial": 0, "weekly": 129, "monthly": 499, "yearly": 5799}

# ─── Models ───────────────────────────────────────────────────────────────────

class UserCreate(BaseModel):
    name: str
    email: str
    password: str

class UserLogin(BaseModel):
    email: str
    password: str

class PhoneSendOTP(BaseModel):
    phone: str

class PhoneVerifyOTP(BaseModel):
    phone: str
    otp: str
    name: Optional[str] = None

class SubscriptionUpdate(BaseModel):
    subscription_type: str   # trial, weekly, monthly, yearly
    duration_days: Optional[int] = None  # override default if set

class BackupSettings(BaseModel):
    backup_email: str
    backup_frequency: str  # off / daily / weekly / monthly

class PartyCreate(BaseModel):
    name: str
    mobile: str = ""
    address: str = ""

class PartyUpdate(BaseModel):
    name: Optional[str] = None
    mobile: Optional[str] = None
    address: Optional[str] = None

class LedgerEntryCreate(BaseModel):
    date: str
    jama: float = 0.0
    naam: float = 0.0
    narration: str = ""
    counterparty_id: Optional[str] = None

class LedgerEntryUpdate(BaseModel):
    date: Optional[str] = None
    jama: Optional[float] = None
    naam: Optional[float] = None
    narration: Optional[str] = None

# ─── Auth Utils ───────────────────────────────────────────────────────────────

JWT_ALGORITHM = "HS256"

def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode("utf-8"), salt).decode("utf-8")

def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))

def create_access_token(user_id: str, email: str) -> str:
    payload = {"sub": user_id, "email": email, "exp": datetime.now(timezone.utc) + timedelta(minutes=30), "type": "access"}
    return jwt.encode(payload, os.environ["JWT_SECRET"], algorithm=JWT_ALGORITHM)

def create_refresh_token(user_id: str) -> str:
    payload = {"sub": user_id, "exp": datetime.now(timezone.utc) + timedelta(days=7), "type": "refresh"}
    return jwt.encode(payload, os.environ["JWT_SECRET"], algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, os.environ["JWT_SECRET"], algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        user["_id"] = str(user["_id"])
        user.pop("password_hash", None)
        # Add subscription info
        sub_type = user.get("subscription_type", "trial")
        sub_expires = user.get("subscription_expires_at")
        now = datetime.now(timezone.utc)
        if sub_expires:
            if hasattr(sub_expires, 'tzinfo') and sub_expires.tzinfo is None:
                sub_expires = sub_expires.replace(tzinfo=timezone.utc)
            is_active = now <= sub_expires
            days_rem = max(0, (sub_expires - now).days)
        else:
            is_active = True
            days_rem = 15
        user["subscription"] = {
            "type": sub_type,
            "expires_at": sub_expires.isoformat() if sub_expires else None,
            "is_active": is_active or user.get("role") in ["admin", "superadmin"],
            "days_remaining": days_rem,
        }
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# ─── Balance Utils ────────────────────────────────────────────────────────────

async def recalculate_balances(party_id: str) -> float:
    entries = await db.ledger_entries.find(
        {"party_id": party_id, "is_deleted": {"$ne": True}},
        sort=[("date", 1), ("created_at", 1)]
    ).to_list(None)
    running = 0.0
    for e in entries:
        # NEW formula: balance = prev + naam (credit/lena) - jama (debit/dena)
        # positive balance = Lena Hai (receivable), negative = Dena Hai (payable)
        running = running + e.get("naam", 0.0) - e.get("jama", 0.0)
        await db.ledger_entries.update_one(
            {"_id": e["_id"]}, {"$set": {"balance": round(running, 2)}}
        )
    return round(running, 2)

async def get_party_balance(party_id: str) -> float:
    last = await db.ledger_entries.find_one(
        {"party_id": party_id, "is_deleted": {"$ne": True}},
        sort=[("date", -1), ("created_at", -1)]
    )
    return last.get("balance", 0.0) if last else 0.0

def format_entry(e: dict) -> dict:
    return {
        "id": str(e["_id"]),
        "party_id": e.get("party_id", ""),
        "counterparty_id": e.get("counterparty_id"),
        "counterparty_name": e.get("counterparty_name", ""),   # populated by caller
        "transaction_id": e.get("transaction_id"),
        "date": e.get("date", ""),
        "jama": e.get("jama", 0.0),
        "naam": e.get("naam", 0.0),
        "narration": e.get("narration", ""),
        "balance": e.get("balance", 0.0),
        "is_locked": e.get("is_locked", False),
        "tally_id": e.get("tally_id"),
        "created_at": str(e.get("created_at", "")),
    }

# ─── Startup ──────────────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup_event():
    # Drop old email index and recreate as sparse (supports phone-only users)
    try:
        await db.users.drop_index("email_1")
    except Exception:
        pass
    await db.users.create_index("email", unique=True, sparse=True)
    await db.users.create_index("phone", unique=True, sparse=True)
    await db.parties.create_index([("user_id", 1), ("is_deleted", 1)])
    await db.ledger_entries.create_index([("party_id", 1), ("date", 1)])
    await db.ledger_entries.create_index("transaction_id")
    await db.login_attempts.create_index("identifier")
    # Recalculate all existing balances with updated formula
    all_parties = await db.parties.find({"is_deleted": {"$ne": True}}).to_list(None)
    for p in all_parties:
        await recalculate_balances(str(p["_id"]))

    now = datetime.now(timezone.utc)
    # Seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@khaata.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing_admin = await db.users.find_one({"email": admin_email})
    if existing_admin is None:
        await db.users.insert_one({
            "email": admin_email, "password_hash": hash_password(admin_password),
            "name": "Admin", "role": "admin", "created_at": now,
            "subscription_type": "yearly", "subscription_expires_at": now + timedelta(days=3650),
        })
    elif not verify_password(admin_password, existing_admin.get("password_hash", "")):
        await db.users.update_one({"email": admin_email}, {"$set": {"password_hash": hash_password(admin_password)}})

    # Start APScheduler for scheduled backups
    try:
        from apscheduler.schedulers.asyncio import AsyncIOScheduler
        scheduler = AsyncIOScheduler(timezone="UTC")
        scheduler.add_job(_scheduled_backup_task, "interval", hours=1, id="backup_task")
        scheduler.start()
        app.state.scheduler = scheduler
        logging.info("APScheduler started — backup task runs every hour")
    except Exception as e:
        logging.warning(f"APScheduler failed to start: {e}")
    sa_email = os.environ.get("SUPERADMIN_EMAIL", "superadmin@poketbook.in")
    sa_password = os.environ.get("SUPERADMIN_PASSWORD", "PoketBook@Super2024")
    existing_sa = await db.users.find_one({"email": sa_email})
    if existing_sa is None:
        await db.users.insert_one({
            "email": sa_email, "password_hash": hash_password(sa_password),
            "name": "Super Admin", "role": "superadmin", "created_at": now,
            "subscription_type": "yearly", "subscription_expires_at": now + timedelta(days=3650),
        })
    elif not verify_password(sa_password, existing_sa.get("password_hash", "")):
        await db.users.update_one({"email": sa_email}, {"$set": {"password_hash": hash_password(sa_password)}})

# ─── Auth Routes ──────────────────────────────────────────────────────────────

@api_router.post("/auth/register")
async def register(data: UserCreate, response: Response):
    existing = await db.users.find_one({"email": data.email.lower()})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    now = datetime.now(timezone.utc)
    result = await db.users.insert_one({
        "email": data.email.lower(),
        "password_hash": hash_password(data.password),
        "name": data.name,
        "role": "user",
        "created_at": now,
        # 7-day FREE trial on registration
        "subscription_type": "trial",
        "subscription_started_at": now,
        "subscription_expires_at": now + timedelta(days=7),
        "subscription_is_active": True,
    })
    user_id = str(result.inserted_id)
    access_token = create_access_token(user_id, data.email.lower())
    refresh_tok = create_refresh_token(user_id)
    response.set_cookie("access_token", access_token, httponly=True, secure=False, samesite="lax", max_age=1800, path="/")
    response.set_cookie("refresh_token", refresh_tok, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    return {"id": user_id, "email": data.email.lower(), "name": data.name, "role": "user",
            "access_token": access_token, "refresh_token": refresh_tok}

@api_router.post("/auth/login")
async def login(data: UserLogin, response: Response, request: Request):
    email = data.email.lower()
    ip = request.client.host if request.client else "unknown"
    identifier = f"{ip}:{email}"
    attempt = await db.login_attempts.find_one({"identifier": identifier})
    if attempt and attempt.get("count", 0) >= 5:
        last = attempt.get("last_attempt")
        if last:
            last_aware = last if last.tzinfo else last.replace(tzinfo=timezone.utc)
            if datetime.now(timezone.utc) - last_aware < timedelta(minutes=15):
                raise HTTPException(status_code=429, detail="Too many failed attempts. 15 minute baad try karein.")
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(data.password, user.get("password_hash", "")):
        await db.login_attempts.update_one(
            {"identifier": identifier},
            {"$inc": {"count": 1}, "$set": {"last_attempt": datetime.now(timezone.utc)}},
            upsert=True,
        )
        raise HTTPException(status_code=401, detail="Invalid email or password")
    await db.login_attempts.delete_one({"identifier": identifier})
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, email)
    refresh_tok = create_refresh_token(user_id)
    response.set_cookie("access_token", access_token, httponly=True, secure=False, samesite="lax", max_age=1800, path="/")
    response.set_cookie("refresh_token", refresh_tok, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    return {"id": user_id, "email": email, "name": user.get("name"), "role": user.get("role", "user"),
            "access_token": access_token, "refresh_token": refresh_tok}

@api_router.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"message": "Logged out"}

@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    return current_user

@api_router.post("/auth/refresh")
async def refresh_token_endpoint(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        try:
            body = await request.json()
            token = body.get("refresh_token")
        except Exception:
            pass
    if not token:
        raise HTTPException(status_code=401, detail="No refresh token")
    try:
        payload = jwt.decode(token, os.environ["JWT_SECRET"], algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user = await db.users.find_one({"_id": ObjectId(payload["sub"])})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        access_token = create_access_token(str(user["_id"]), user["email"])
        response.set_cookie("access_token", access_token, httponly=True, secure=False, samesite="lax", max_age=1800, path="/")
        return {"message": "Token refreshed", "access_token": access_token}
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid refresh token")

# ─── Party Routes ─────────────────────────────────────────────────────────────

@api_router.get("/parties")
async def get_parties(current_user: dict = Depends(get_current_user)):
    parties = await db.parties.find(
        {"user_id": current_user["_id"], "is_deleted": {"$ne": True}}
    ).to_list(None)
    result = []
    for p in parties:
        pid = str(p["_id"])
        bal = await get_party_balance(pid)
        unlocked = await db.ledger_entries.count_documents(
            {"party_id": pid, "is_locked": {"$ne": True}, "is_deleted": {"$ne": True}}
        )
        balance_zero = abs(round(bal, 2)) <= 0.01
        result.append({
            "id": pid, "name": p["name"],
            "mobile": p.get("mobile", ""), "address": p.get("address", ""),
            "current_balance": bal, "created_at": str(p.get("created_at", "")),
            "has_unlocked_entries": unlocked > 0,
            "is_deletable": balance_zero and unlocked == 0,
            "delete_reason": (
                "" if (balance_zero and unlocked == 0)
                else ("Balance zero nahi hai" if not balance_zero else f"{unlocked} unlocked entries hain")
            ),
        })
    return result

@api_router.post("/parties")
async def create_party(data: PartyCreate, current_user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    result = await db.parties.insert_one({
        "name": data.name, "mobile": data.mobile, "address": data.address,
        "user_id": current_user["_id"], "is_deleted": False,
        "created_at": now, "updated_at": now,
    })
    return {"id": str(result.inserted_id), "name": data.name, "mobile": data.mobile, "address": data.address, "current_balance": 0.0}

@api_router.get("/parties/{party_id}")
async def get_party(party_id: str, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"], "is_deleted": {"$ne": True}})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    bal = await get_party_balance(party_id)
    return {"id": party_id, "name": p["name"], "mobile": p.get("mobile", ""), "address": p.get("address", ""), "current_balance": bal}

@api_router.put("/parties/{party_id}")
async def update_party(party_id: str, data: PartyUpdate, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"], "is_deleted": {"$ne": True}})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    update = {k: v for k, v in data.model_dump().items() if v is not None}
    update["updated_at"] = datetime.now(timezone.utc)
    await db.parties.update_one({"_id": ObjectId(party_id)}, {"$set": update})
    updated = await db.parties.find_one({"_id": ObjectId(party_id)})
    bal = await get_party_balance(party_id)
    return {"id": party_id, "name": updated["name"], "mobile": updated.get("mobile", ""), "address": updated.get("address", ""), "current_balance": bal}

@api_router.delete("/parties/{party_id}")
async def delete_party(party_id: str, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    # Rule 1: Balance must be zero (within 1 paisa tolerance)
    balance = await get_party_balance(party_id)
    if abs(round(balance, 2)) > 0.01:
        bal_text = f"₹{abs(balance):.2f} {'Dena Hai' if balance > 0 else 'Lena Hai'}"
        raise HTTPException(status_code=400, detail=f"Balance clear karein pehle. Current: {bal_text}")
    # Rule 2: No unlocked entries
    unlocked = await db.ledger_entries.count_documents(
        {"party_id": party_id, "is_locked": {"$ne": True}, "is_deleted": {"$ne": True}}
    )
    if unlocked > 0:
        raise HTTPException(status_code=400, detail=f"{unlocked} unlocked entries hain. Pehle Tally/Lock karein.")
    # Soft delete
    await db.parties.update_one({"_id": ObjectId(party_id)}, {"$set": {"is_deleted": True, "updated_at": datetime.now(timezone.utc)}})
    return {"message": "Party deleted successfully"}

# ─── Ledger Routes ────────────────────────────────────────────────────────────

@api_router.get("/ledger/{party_id}/entries")
async def get_ledger_entries(party_id: str, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    entries = await db.ledger_entries.find(
        {"party_id": party_id, "is_deleted": {"$ne": True}},
        sort=[("date", 1), ("created_at", 1)]    # ASC — oldest first (top), newest at bottom
    ).to_list(None)
    # Batch-load counterparty names
    cp_ids = list({e["counterparty_id"] for e in entries if e.get("counterparty_id")})
    cp_map = {}
    if cp_ids:
        cps = await db.parties.find({"_id": {"$in": [ObjectId(cid) for cid in cp_ids]}}).to_list(None)
        cp_map = {str(c["_id"]): c["name"] for c in cps}
    formatted = []
    for e in entries:
        fe = format_entry(e)
        fe["counterparty_name"] = cp_map.get(e.get("counterparty_id", ""), "")
        formatted.append(fe)
    # current_balance = most recent entry (last in ASC order)
    current_balance = formatted[-1]["balance"] if formatted else 0.0
    return {
        "party": {"id": party_id, "name": p["name"], "mobile": p.get("mobile", ""), "address": p.get("address", "")},
        "entries": formatted,
        "current_balance": current_balance,
    }

@api_router.post("/ledger/{party_id}/entries")
async def add_ledger_entry(party_id: str, data: LedgerEntryCreate, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    # Validate counterparty if provided
    cp = None
    if data.counterparty_id and data.counterparty_id != party_id:
        cp = await db.parties.find_one({"_id": ObjectId(data.counterparty_id), "user_id": current_user["_id"], "is_deleted": {"$ne": True}})
        if not cp:
            raise HTTPException(status_code=404, detail="Counterparty not found")
    now = datetime.now(timezone.utc)
    transaction_id = str(uuid_lib.uuid4())
    # Entry A — Header party's ledger
    result_a = await db.ledger_entries.insert_one({
        "party_id": party_id, "counterparty_id": data.counterparty_id or None,
        "transaction_id": transaction_id, "date": data.date,
        "jama": data.jama, "naam": data.naam, "narration": data.narration,
        "balance": 0.0, "is_locked": False, "tally_id": None,
        "is_deleted": False, "created_at": now, "updated_at": now,
    })
    await recalculate_balances(party_id)
    # Entry B — Counterparty's ledger (mirrored: naam↔jama swapped)
    if cp:
        await db.ledger_entries.insert_one({
            "party_id": data.counterparty_id, "counterparty_id": party_id,
            "transaction_id": transaction_id, "date": data.date,
            "jama": data.naam,   # A's naam becomes B's jama
            "naam": data.jama,   # A's jama becomes B's naam
            "narration": data.narration,
            "balance": 0.0, "is_locked": False, "tally_id": None,
            "is_deleted": False, "created_at": now, "updated_at": now,
        })
        await recalculate_balances(data.counterparty_id)
    updated_a = await db.ledger_entries.find_one({"_id": result_a.inserted_id})
    fe = format_entry(updated_a)
    fe["counterparty_name"] = cp["name"] if cp else ""
    return fe

@api_router.put("/ledger/{party_id}/entries/{entry_id}")
async def update_ledger_entry(party_id: str, entry_id: str, data: LedgerEntryUpdate, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    e = await db.ledger_entries.find_one({"_id": ObjectId(entry_id), "party_id": party_id})
    if not e:
        raise HTTPException(status_code=404, detail="Entry not found")
    if e.get("is_locked"):
        raise HTTPException(status_code=403, detail="Locked entry edit nahi ho sakti")
    update_fields = {k: v for k, v in data.model_dump().items() if v is not None}
    update_fields["updated_at"] = datetime.now(timezone.utc)
    await db.ledger_entries.update_one({"_id": ObjectId(entry_id)}, {"$set": update_fields})
    await recalculate_balances(party_id)
    # Update mirrored entry (swap naam/jama)
    txn_id = e.get("transaction_id")
    if txn_id:
        mirror = await db.ledger_entries.find_one(
            {"transaction_id": txn_id, "party_id": {"$ne": party_id}, "is_deleted": {"$ne": True}}
        )
        if mirror and not mirror.get("is_locked"):
            mirror_update = {"updated_at": datetime.now(timezone.utc)}
            if "naam" in update_fields:
                mirror_update["jama"] = update_fields["naam"]
            if "jama" in update_fields:
                mirror_update["naam"] = update_fields["jama"]
            if "date" in update_fields:
                mirror_update["date"] = update_fields["date"]
            if "narration" in update_fields:
                mirror_update["narration"] = update_fields["narration"]
            await db.ledger_entries.update_one({"_id": mirror["_id"]}, {"$set": mirror_update})
            await recalculate_balances(mirror["party_id"])
    updated = await db.ledger_entries.find_one({"_id": ObjectId(entry_id)})
    fe = format_entry(updated)
    if e.get("counterparty_id"):
        cp = await db.parties.find_one({"_id": ObjectId(e["counterparty_id"])})
        fe["counterparty_name"] = cp["name"] if cp else ""
    return fe

@api_router.delete("/ledger/{party_id}/entries/{entry_id}")
async def delete_ledger_entry(party_id: str, entry_id: str, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    e = await db.ledger_entries.find_one({"_id": ObjectId(entry_id), "party_id": party_id})
    if not e:
        raise HTTPException(status_code=404, detail="Entry not found")
    if e.get("is_locked"):
        raise HTTPException(status_code=403, detail="Locked entry delete nahi ho sakti")
    now = datetime.now(timezone.utc)
    await db.ledger_entries.update_one({"_id": ObjectId(entry_id)}, {"$set": {"is_deleted": True, "updated_at": now}})
    await recalculate_balances(party_id)
    # Also delete the mirrored entry (same transaction_id)
    txn_id = e.get("transaction_id")
    if txn_id:
        mirror = await db.ledger_entries.find_one(
            {"transaction_id": txn_id, "party_id": {"$ne": party_id}, "is_deleted": {"$ne": True}}
        )
        if mirror and not mirror.get("is_locked"):
            await db.ledger_entries.update_one({"_id": mirror["_id"]}, {"$set": {"is_deleted": True, "updated_at": now}})
            await recalculate_balances(mirror["party_id"])
    return {"message": "Entry deleted successfully (both sides)"}

@api_router.post("/ledger/{party_id}/tally")
async def lock_party_entries(party_id: str, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    unlocked = await db.ledger_entries.find(
        {"party_id": party_id, "is_locked": {"$ne": True}, "is_deleted": {"$ne": True}}
    ).to_list(None)
    if not unlocked:
        raise HTTPException(status_code=400, detail="Koi unlocked entry nahi hai")
    closing_balance = await get_party_balance(party_id)
    now = datetime.now(timezone.utc)
    tally_result = await db.tallies.insert_one({
        "party_id": party_id, "user_id": current_user["_id"],
        "closing_balance": closing_balance, "timestamp": now, "type": "party",
    })
    tally_id = str(tally_result.inserted_id)
    # Lock this party's entries
    await db.ledger_entries.update_many(
        {"party_id": party_id, "is_locked": {"$ne": True}, "is_deleted": {"$ne": True}},
        {"$set": {"is_locked": True, "tally_id": tally_id}},
    )
    # Lock mirrored entries (same transaction_ids) in counterparties
    txn_ids = [e["transaction_id"] for e in unlocked if e.get("transaction_id")]
    if txn_ids:
        await db.ledger_entries.update_many(
            {"transaction_id": {"$in": txn_ids}, "party_id": {"$ne": party_id},
             "is_locked": {"$ne": True}, "is_deleted": {"$ne": True}},
            {"$set": {"is_locked": True, "tally_id": tally_id}},
        )
    return {"tally_id": tally_id, "closing_balance": closing_balance, "locked_count": len(unlocked), "timestamp": now.isoformat()}

@api_router.get("/ledger/{party_id}/tallies")
async def get_party_tallies(party_id: str, current_user: dict = Depends(get_current_user)):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    tallies = await db.tallies.find({"party_id": party_id}, sort=[("timestamp", -1)]).to_list(None)
    return [{"id": str(t["_id"]), "closing_balance": t.get("closing_balance", 0), "timestamp": str(t.get("timestamp", ""))} for t in tallies]

# ─── Balance Sheet ────────────────────────────────────────────────────────────

@api_router.get("/balance-sheet")
async def get_balance_sheet(current_user: dict = Depends(get_current_user)):
    parties = await db.parties.find({"user_id": current_user["_id"], "is_deleted": {"$ne": True}}).to_list(None)
    lena_hai, dena_hai = [], []
    for p in parties:
        pid = str(p["_id"])
        bal = await get_party_balance(pid)
        entry = {"id": pid, "name": p["name"], "mobile": p.get("mobile", ""), "balance": bal}
        # NAAM (credit) = DENA for party → positive balance → dena_hai (LEFT, BLUE)
        # JAMA (debit)  = LENA for party → negative balance → lena_hai (RIGHT, RED)
        if bal > 0:
            dena_hai.append({**entry, "amount": abs(bal)})   # positive = DENA (gave credit, will pay)
        elif bal < 0:
            lena_hai.append({**entry, "amount": abs(bal)})   # negative = LENA (received credit, will receive)
    total_dena = sum(x["amount"] for x in dena_hai)   # total of DENA parties
    total_lena = sum(x["amount"] for x in lena_hai)   # total of LENA parties
    return {
        "lena_hai": sorted(lena_hai, key=lambda x: x["amount"], reverse=True),
        "dena_hai": sorted(dena_hai, key=lambda x: x["amount"], reverse=True),
        "total_receivable": round(total_dena, 2),   # we will receive from DENA parties
        "total_payable": round(total_lena, 2),      # we will pay to LENA parties
        # net: negative = net LENA (we receive more) → formatBalance(negative) = Lena Hai RED ✓
        "net_balance": round(total_lena - total_dena, 2),
    }

# ─── Export Helper Functions ──────────────────────────────────────────────────

def fmt_inr(amount: float) -> str:
    return f"Rs.{abs(float(amount)):,.2f}"

def _format_balance_text(bal: float) -> str:
    """Format a balance value as human-readable Lena/Dena text."""
    if bal > 0:
        return f"{fmt_inr(bal)} Lena Hai"
    if bal < 0:
        return f"{fmt_inr(bal)} Dena Hai"
    return "Settled"

def _build_date_query(start_date: Optional[str], end_date: Optional[str]) -> dict:
    """Build a MongoDB date range sub-query."""
    if not start_date and not end_date:
        return {}
    date_q: dict = {}
    if start_date:
        date_q["$gte"] = start_date
    if end_date:
        date_q["$lte"] = end_date
    return {"date": date_q}

async def _load_entries_with_cp_names(party_id: str, extra_query: dict) -> tuple[list, dict]:
    """Fetch ledger entries and return (entries, counterparty_name_map)."""
    query = {"party_id": party_id, "is_deleted": {"$ne": True}, **extra_query}
    entries = await db.ledger_entries.find(
        query, sort=[("date", 1), ("created_at", 1)]
    ).to_list(None)
    cp_ids = list({e["counterparty_id"] for e in entries if e.get("counterparty_id")})
    cp_map: dict = {}
    if cp_ids:
        cps = await db.parties.find({"_id": {"$in": [ObjectId(c) for c in cp_ids]}}).to_list(None)
        cp_map = {str(c["_id"]): c["name"] for c in cps}
    return entries, cp_map

def _build_ledger_pdf_table_data(entries: list, cp_map: dict) -> list:
    """Convert ledger entries to a list of rows for the PDF table."""
    headers = ["Date", "Party Name", "Credit (नाम)", "Debit (जमा)", "Narration", "Balance", "Status"]
    rows = [headers]
    for e in entries:
        rows.append([
            e.get("date", ""),
            cp_map.get(e.get("counterparty_id", ""), "—"),
            fmt_inr(e.get("naam", 0)) if e.get("naam", 0) > 0 else "-",
            fmt_inr(e.get("jama", 0)) if e.get("jama", 0) > 0 else "-",
            str(e.get("narration", "") or "")[:35],
            _format_balance_text(e.get("balance", 0)),
            "Locked" if e.get("is_locked") else "Open",
        ])
    return rows

def _build_ledger_pdf_table_style():
    """Return the TableStyle for the ledger PDF table."""
    from reportlab.platypus import TableStyle
    from reportlab.lib import colors
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8B4513")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#FFE8CC"), colors.HexColor("#FFDAB0")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D6D3D1")),
        ("ALIGN", (2, 1), (3, -1), "RIGHT"),
        ("ALIGN", (5, 1), (5, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])

def _build_excel_ledger(party_name: str, date_range: str, entries: list, cp_map: dict):
    """Create and populate an openpyxl workbook for a ledger export."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(party_name)[:31]
    ws.append(["Statement", party_name])
    ws.append(["Period", date_range])
    ws.append(["Generated", datetime.now().strftime("%d-%m-%Y %H:%M")])
    ws.append([])
    headers = ["Date", "Party Name", "Credit/Naam", "Debit/Jama", "Narration", "Balance", "Balance Type", "Status"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8B4513")
    for e in entries:
        bal = e.get("balance", 0)
        bal_type = "Lena Hai" if bal > 0 else ("Dena Hai" if bal < 0 else "Settled")
        cp_name = cp_map.get(e.get("counterparty_id", ""), "")
        ws.append([
            e.get("date", ""), cp_name,
            e.get("naam", 0) or "", e.get("jama", 0) or "",
            e.get("narration", "") or "", abs(bal), bal_type,
            "Locked" if e.get("is_locked") else "Open",
        ])
    return wb

# ─── Export Routes (simplified using helpers above) ───────────────────────────

@api_router.get("/export/ledger/{party_id}/pdf")
async def export_ledger_pdf(
    party_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")

    entries, cp_map = await _load_entries_with_cp_names(party_id, _build_date_query(start_date, end_date))
    date_range = f"{start_date or 'All'} to {end_date or 'Today'}"
    styles = getSampleStyleSheet()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    table_data = _build_ledger_pdf_table_data(entries, cp_map)
    t = Table(table_data, colWidths=[60, 80, 80, 80, 160, 110, 50])
    t.setStyle(_build_ledger_pdf_table_style())
    elements = [
        Paragraph(f"Statement: {p['name']}", styles["Title"]),
        Paragraph(f"Period: {date_range}   |   Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]),
        Spacer(1, 12),
        t,
    ]
    doc.build(elements)
    buf.seek(0)
    return Response(content=buf.read(), media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename=Statement_{p['name']}.pdf"})

@api_router.get("/export/ledger/{party_id}/excel")
async def export_ledger_excel(
    party_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")

    entries, cp_map = await _load_entries_with_cp_names(party_id, _build_date_query(start_date, end_date))
    date_range = f"{start_date or 'All'} to {end_date or 'Today'}"
    wb = _build_excel_ledger(p["name"], date_range, entries, cp_map)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=Statement_{p['name']}.xlsx"})
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    p = await db.parties.find_one({"_id": ObjectId(party_id), "user_id": current_user["_id"]})
    if not p:
        raise HTTPException(status_code=404, detail="Party not found")
    query = {"party_id": party_id, "is_deleted": {"$ne": True}}
    if start_date or end_date:
        date_q = {}
        if start_date:
            date_q["$gte"] = start_date
        if end_date:
            date_q["$lte"] = end_date
        query["date"] = date_q
    entries = await db.ledger_entries.find(query, sort=[("date", 1), ("created_at", 1)]).to_list(None)
    # Batch-load counterparty names
    cp_ids = list({e["counterparty_id"] for e in entries if e.get("counterparty_id")})
    cp_map = {}
    if cp_ids:
        cps = await db.parties.find({"_id": {"$in": [ObjectId(cid) for cid in cp_ids]}}).to_list(None)
        cp_map = {str(c["_id"]): c["name"] for c in cps}
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(p["name"])[:31]
    ws.append(["Statement", p["name"]])
    ws.append(["Period", f"{start_date or 'All'} to {end_date or 'Today'}"])
    ws.append(["Generated", datetime.now().strftime("%d-%m-%Y %H:%M")])
    ws.append([])
    headers = ["Date", "Party Name", "Credit/Naam (लेना)", "Debit/Jama (देना)", "Narration", "Balance", "Balance Type", "Status"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="8B4513")
    for e in entries:
        bal = e.get("balance", 0)
        bal_type = "Lena Hai" if bal > 0 else ("Dena Hai" if bal < 0 else "Settled")
        cp_name = cp_map.get(e.get("counterparty_id", ""), "")
        ws.append([e.get("date", ""), cp_name,
                   e.get("naam", 0) or "", e.get("jama", 0) or "",
                   e.get("narration", "") or "", abs(bal), bal_type,
                   "Locked" if e.get("is_locked") else "Open"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=Statement_{p['name']}.xlsx"})

@api_router.get("/export/balance-sheet/pdf")
async def export_bs_pdf(current_user: dict = Depends(get_current_user)):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    data = await get_balance_sheet(current_user)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Balance Sheet", styles["Title"]),
        Paragraph(f"Generated: {datetime.now().strftime('%d-%m-%Y')}", styles["Normal"]),
        Spacer(1, 12),
    ]
    lena, dena = data["lena_hai"], data["dena_hai"]
    max_rows = max(len(lena), len(dena), 1)
    table_data = [["Lena Hai (Receivable)", "Amount", "Dena Hai (Payable)", "Amount"]]
    for i in range(max_rows):
        lena_row = lena[i] if i < len(lena) else {"name": "", "amount": ""}
        dena_row = dena[i] if i < len(dena) else {"name": "", "amount": ""}
        table_data.append([
            lena_row.get("name", ""), fmt_inr(lena_row["amount"]) if lena_row.get("name") else "",
            dena_row.get("name", ""), fmt_inr(dena_row["amount"]) if dena_row.get("name") else "",
        ])
    table_data.append(["TOTAL RECEIVABLE", fmt_inr(data["total_receivable"]), "TOTAL PAYABLE", fmt_inr(data["total_payable"])])
    t = Table(table_data, colWidths=[150, 90, 150, 90])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#15803D")),
        ("BACKGROUND", (2, 0), (3, 0), colors.HexColor("#B91C1C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F5F5F4")]),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"), ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return Response(content=buf.read(), media_type="application/pdf",
                    headers={"Content-Disposition": "attachment; filename=balance_sheet.pdf"})

@api_router.get("/export/balance-sheet/excel")
async def export_bs_excel(current_user: dict = Depends(get_current_user)):
    import openpyxl
    from openpyxl.styles import Font
    data = await get_balance_sheet(current_user)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws.append(["Balance Sheet"])
    ws.append(["Generated", datetime.now().strftime("%d-%m-%Y %H:%M")])
    ws.append([])
    ws.append(["Lena Hai (Receivable)", "Amount", "", "Dena Hai (Payable)", "Amount"])
    for cell in ws[4]:
        cell.font = Font(bold=True)
    max_rows = max(len(data["lena_hai"]), len(data["dena_hai"]), 1)
    for i in range(max_rows):
        lena_row = data["lena_hai"][i] if i < len(data["lena_hai"]) else {}
        dena_row = data["dena_hai"][i] if i < len(data["dena_hai"]) else {}
        ws.append([lena_row.get("name", ""), lena_row.get("amount", "") or "", "", dena_row.get("name", ""), dena_row.get("amount", "") or ""])
    ws.append(["TOTAL RECEIVABLE", data["total_receivable"], "", "TOTAL PAYABLE", data["total_payable"]])
    ws.append(["NET BALANCE", data["net_balance"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=balance_sheet.xlsx"})

# ─── Phone OTP (MSG91 placeholder — activate with MSG91_AUTH_KEY) ─────────────

@api_router.post("/auth/send-otp")
async def send_otp_endpoint(data: PhoneSendOTP):
    phone = data.phone.strip()
    if not phone.startswith("+"):
        phone = "+91" + phone.lstrip("0")
    otp = "".join(secrets.choice(string.digits) for _ in range(6))  # cryptographically secure OTP
    _otp_store[phone] = {"otp": otp, "expires_at": datetime.now(timezone.utc) + timedelta(minutes=5), "attempts": 0}
    msg91_key = os.environ.get("MSG91_AUTH_KEY", "PLACEHOLDER")
    if msg91_key and msg91_key != "PLACEHOLDER":
        try:
            import requests as req
            req.post("https://api.msg91.com/api/v5/otp",
                     params={"authkey": msg91_key, "mobile": phone.lstrip("+"), "otp_length": 6, "otp_expiry": 5},
                     json={"message": f"Your poketbook OTP is {otp}. Valid 5 minutes."})
        except Exception as e:
            logging.error(f"MSG91 error: {e}")
    dev_mode = (msg91_key == "PLACEHOLDER")
    result = {"message": "OTP sent", "phone": phone}
    if dev_mode:
        result["dev_otp"] = otp  # REMOVE IN PRODUCTION
    return result

@api_router.post("/auth/verify-otp")
async def verify_otp_endpoint(data: PhoneVerifyOTP, response: Response):
    phone = data.phone.strip()
    if not phone.startswith("+"):
        phone = "+91" + phone.lstrip("0")
    stored = _otp_store.get(phone)
    if not stored:
        raise HTTPException(400, "OTP not found or expired. Request a new OTP.")
    now = datetime.now(timezone.utc)
    exp = stored["expires_at"]
    if exp.tzinfo is None:
        exp = exp.replace(tzinfo=timezone.utc)
    if now > exp:
        _otp_store.pop(phone, None)
        raise HTTPException(400, "OTP expired. Please request a new OTP.")
    if stored["attempts"] >= 3:
        _otp_store.pop(phone, None)
        raise HTTPException(400, "Too many wrong attempts. Request a new OTP.")
    if stored["otp"] != data.otp:
        _otp_store[phone]["attempts"] += 1
        raise HTTPException(400, f"Invalid OTP. {3 - stored['attempts']} attempts remaining.")
    _otp_store.pop(phone, None)
    user = await db.users.find_one({"phone": phone})
    if not user:
        reg_now = datetime.now(timezone.utc)
        result = await db.users.insert_one({
            "phone": phone, "name": data.name or f"User {phone[-4:]}",
            "email": None, "password_hash": "", "role": "user", "created_at": reg_now,
            "subscription_type": "trial",
            "subscription_started_at": reg_now,
            "subscription_expires_at": reg_now + timedelta(days=7),
            "subscription_is_active": True,
        })
        user = await db.users.find_one({"_id": result.inserted_id})
    user_id = str(user["_id"])
    access_token = create_access_token(user_id, phone)
    refresh_tok = create_refresh_token(user_id)
    response.set_cookie("access_token", access_token, httponly=True, secure=False, samesite="lax", max_age=1800, path="/")
    response.set_cookie("refresh_token", refresh_tok, httponly=True, secure=False, samesite="lax", max_age=604800, path="/")
    return {"id": user_id, "name": user.get("name"), "phone": phone,
            "access_token": access_token, "refresh_token": refresh_tok}

# ─── Superadmin Panel ─────────────────────────────────────────────────────────

def _require_superadmin(current_user: dict):
    if current_user.get("role") not in ["superadmin", "admin"]:
        raise HTTPException(status_code=403, detail="Superadmin access required")

@api_router.get("/superadmin/stats")
async def sa_stats(current_user: dict = Depends(get_current_user)):
    _require_superadmin(current_user)
    now = datetime.now(timezone.utc)
    total = await db.users.count_documents({"role": {"$nin": ["superadmin", "admin"]}})
    active = await db.users.count_documents({"role": {"$nin": ["superadmin","admin"]}, "subscription_expires_at": {"$gt": now}})
    trial = await db.users.count_documents({"role": {"$nin": ["superadmin","admin"]}, "subscription_type": "trial", "subscription_expires_at": {"$gt": now}})
    return {"total_users": total, "active_subscriptions": active, "trial_users": trial, "expired": total - active}

@api_router.get("/superadmin/users")
async def sa_get_users(current_user: dict = Depends(get_current_user)):
    _require_superadmin(current_user)
    users = await db.users.find({"role": {"$nin": ["superadmin", "admin"]}}).sort("created_at", -1).to_list(None)
    now = datetime.now(timezone.utc)
    result = []
    for u in users:
        sub_exp = u.get("subscription_expires_at")
        if sub_exp and sub_exp.tzinfo is None:
            sub_exp = sub_exp.replace(tzinfo=timezone.utc)
        is_active = bool(sub_exp and now <= sub_exp)
        days_rem = max(0, (sub_exp - now).days) if sub_exp and is_active else 0
        result.append({
            "id": str(u["_id"]), "name": u.get("name", ""), "email": u.get("email") or "",
            "phone": u.get("phone", ""), "created_at": str(u.get("created_at", "")),
            "subscription_type": u.get("subscription_type", "trial"),
            "subscription_expires_at": str(sub_exp) if sub_exp else None,
            "subscription_active": is_active, "days_remaining": days_rem,
        })
    return result

@api_router.post("/superadmin/users/{user_id}/subscription")
async def sa_update_subscription(user_id: str, data: SubscriptionUpdate, current_user: dict = Depends(get_current_user)):
    _require_superadmin(current_user)
    days = data.duration_days or SUBSCRIPTION_DAYS.get(data.subscription_type, 30)
    now = datetime.now(timezone.utc)
    await db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {
        "subscription_type": data.subscription_type,
        "subscription_started_at": now,
        "subscription_expires_at": now + timedelta(days=days),
        "subscription_is_active": True,
    }})
    return {"message": f"Subscription set to {data.subscription_type} for {days} days"}

# ─── CSV/Google Sheets Backup ─────────────────────────────────────────────────

@api_router.get("/export/csv-backup")
async def export_csv_backup(current_user: dict = Depends(get_current_user)):
    """Export all user data as CSV — ready to import into Google Sheets"""
    import csv
    parties = await db.parties.find({"user_id": current_user["_id"], "is_deleted": {"$ne": True}}).to_list(None)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["=== PARTIES ==="])
    writer.writerow(["ID", "Name", "Mobile", "Address", "Current Balance"])
    for p in parties:
        pid = str(p["_id"])
        bal = await get_party_balance(pid)
        writer.writerow([pid, p["name"], p.get("mobile",""), p.get("address",""), bal])
    writer.writerow([])
    writer.writerow(["=== LEDGER ENTRIES ==="])
    writer.writerow(["Date", "Party", "Counterparty", "Credit/Naam", "Debit/Jama", "Narration", "Balance", "Locked"])
    for p in parties:
        pid = str(p["_id"])
        entries = await db.ledger_entries.find({"party_id": pid, "is_deleted": {"$ne": True}}, sort=[("date",1)]).to_list(None)
        cp_ids = list({e["counterparty_id"] for e in entries if e.get("counterparty_id")})
        cp_map = {}
        if cp_ids:
            cps = await db.parties.find({"_id": {"$in": [ObjectId(c) for c in cp_ids]}}).to_list(None)
            cp_map = {str(c["_id"]): c["name"] for c in cps}
        for e in entries:
            writer.writerow([e.get("date",""), p["name"], cp_map.get(e.get("counterparty_id",""),""),
                             e.get("naam",0), e.get("jama",0), e.get("narration",""),
                             e.get("balance",0), "Yes" if e.get("is_locked") else "No"])
    csv_content = buf.getvalue().encode("utf-8-sig")  # UTF-8 BOM for Excel/Sheets compatibility
    return Response(content=csv_content, media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=poketbook_backup_{datetime.now().strftime('%Y%m%d')}.csv"})

# ─── Google Sheets OAuth (Placeholder — activate with GOOGLE_CLIENT_SECRET) ───

@api_router.get("/oauth/sheets/connect")
async def sheets_connect_url(current_user: dict = Depends(get_current_user)):
    """Returns Google OAuth URL for Sheets access"""
    client_id = os.environ.get("GOOGLE_CLIENT_ID", "")
    client_secret = os.environ.get("GOOGLE_CLIENT_SECRET", "PLACEHOLDER")
    redirect_uri = os.environ.get("GOOGLE_REDIRECT_URI", "")
    if client_secret == "PLACEHOLDER":
        return {"error": "Google Sheets integration not configured.", "configured": False}
    try:
        from google_auth_oauthlib.flow import Flow
        flow = Flow.from_client_config(
            {"web": {"client_id": client_id, "client_secret": client_secret,
                     "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                     "token_uri": "https://oauth2.googleapis.com/token"}},
            scopes=GOOGLE_SCOPES,
            redirect_uri=redirect_uri
        )
        url, state = flow.authorization_url(access_type="offline", prompt="consent")
        await db.oauth_states.insert_one({
            "state": state, "user_id": current_user["_id"],
            "expires_at": datetime.now(timezone.utc) + timedelta(minutes=10)
        })
        return {"url": url, "configured": True}
    except Exception as e:
        logging.error(f"Sheets OAuth error: {e}")
        return {"error": str(e), "configured": False}

@api_router.get("/oauth/sheets/callback")
async def sheets_oauth_callback(code: str = None, state: str = None, error: str = None):
    """Handle OAuth callback from Google — store tokens and redirect to app"""
    from fastapi.responses import RedirectResponse
    if error or not code:
        return RedirectResponse(url="/?sheets=error")
    try:
        state_doc = await db.oauth_states.find_one({"state": state})
        if not state_doc:
            return RedirectResponse(url="/?sheets=error&reason=invalid_state")
        user_id = state_doc["user_id"]
        await db.oauth_states.delete_one({"state": state})
        client_id = os.environ.get("GOOGLE_CLIENT_ID")
        client_secret = os.environ.get("GOOGLE_CLIENT_SECRET")
        redirect_uri = os.environ.get("GOOGLE_REDIRECT_URI")
        from google_auth_oauthlib.flow import Flow
        flow = Flow.from_client_config(
            {"web": {"client_id": client_id, "client_secret": client_secret,
                     "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                     "token_uri": "https://oauth2.googleapis.com/token"}},
            scopes=GOOGLE_SCOPES,
            redirect_uri=redirect_uri
        )
        flow.fetch_token(code=code)
        creds = flow.credentials
        await db.users.update_one({"_id": ObjectId(user_id) if not isinstance(user_id, ObjectId) else user_id}, {"$set": {
            "google_access_token": creds.token,
            "google_refresh_token": creds.refresh_token,
            "google_token_expiry": creds.expiry.isoformat() if creds.expiry else None,
            "google_sheets_connected": True,
        }})
        frontend_url = os.environ.get("FRONTEND_URL", "http://localhost:3000")
        return RedirectResponse(url=f"{frontend_url}/export?sheets=connected")
    except Exception as e:
        logging.error(f"OAuth callback error: {e}")
        return RedirectResponse(url="/?sheets=error")

@api_router.post("/export/google-sheets-backup")
async def backup_to_google_sheets(current_user: dict = Depends(get_current_user)):
    """Write all user data to Google Sheets and return the sheet URL"""
    user = await db.users.find_one({"_id": ObjectId(current_user["_id"])})
    if not user or not user.get("google_sheets_connected"):
        raise HTTPException(400, "Google Sheets not connected. Please connect first.")
    access_token = user.get("google_access_token")
    refresh_token = user.get("google_refresh_token")
    if not access_token and not refresh_token:
        raise HTTPException(400, "Google credentials missing. Please reconnect.")
    try:
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build
        creds = Credentials(
            token=access_token,
            refresh_token=refresh_token,
            token_uri="https://oauth2.googleapis.com/token",
            client_id=os.environ.get("GOOGLE_CLIENT_ID"),
            client_secret=os.environ.get("GOOGLE_CLIENT_SECRET"),
        )
        sheets_svc = build("sheets", "v4", credentials=creds)
        existing_sheet_id = user.get("google_sheet_id")
        if existing_sheet_id:
            try:
                sheets_svc.spreadsheets().get(spreadsheetId=existing_sheet_id).execute()
                sheet_id = existing_sheet_id
            except Exception:
                existing_sheet_id = None
        if not existing_sheet_id:
            spreadsheet = {
                "properties": {"title": f"PoketBook Backup — {user.get('name','User')}"},
                "sheets": [
                    {"properties": {"title": "Parties", "index": 0}},
                    {"properties": {"title": "Ledger Entries", "index": 1}},
                ]
            }
            result = sheets_svc.spreadsheets().create(body=spreadsheet).execute()
            sheet_id = result["spreadsheetId"]
            await db.users.update_one({"_id": ObjectId(current_user["_id"])}, {"$set": {"google_sheet_id": sheet_id}})
        # Fetch all data
        uid = current_user["_id"]
        parties = await db.parties.find({"user_id": uid, "is_deleted": {"$ne": True}}).to_list(None)
        # Parties sheet
        parties_rows = [["ID", "Name", "Mobile", "Address", "Current Balance", "Created At"]]
        for p in parties:
            pid = str(p["_id"])
            bal = await get_party_balance(pid)
            parties_rows.append([pid, p["name"], p.get("mobile",""), p.get("address",""), bal, str(p.get("created_at",""))])
        # Ledger entries
        entries_rows = [["Date", "Time", "Party", "Counterparty", "Credit (Naam)", "Debit (Jama)", "Narration", "Balance", "Balance Type", "Locked"]]
        cp_cache = {}
        for p in parties:
            pid = str(p["_id"])
            entries = await db.ledger_entries.find({"party_id": pid, "is_deleted": {"$ne": True}}, sort=[("date",1),("created_at",1)]).to_list(None)
            for e in entries:
                cp_id = e.get("counterparty_id","")
                if cp_id and cp_id not in cp_cache:
                    cp_doc = await db.parties.find_one({"_id": ObjectId(cp_id)})
                    cp_cache[cp_id] = cp_doc["name"] if cp_doc else ""
                bal = e.get("balance", 0)
                bal_type = "Dena Hai" if bal > 0 else ("Lena Hai" if bal < 0 else "Settled")
                created = str(e.get("created_at",""))
                time_str = ""
                try:
                    from datetime import datetime as dt2
                    d = dt2.fromisoformat(created.replace("+00:00",""))
                    time_str = (d + timedelta(hours=5, minutes=30)).strftime("%I:%M %p IST")
                except Exception:
                    pass
                entries_rows.append([e.get("date",""), time_str, p["name"], cp_cache.get(cp_id,""),
                                     e.get("naam",0) or "", e.get("jama",0) or "",
                                     e.get("narration","") or "", abs(bal), bal_type,
                                     "Yes" if e.get("is_locked") else "No"])
        # Clear and write sheets
        sheets_svc.spreadsheets().values().clear(spreadsheetId=sheet_id, range="Parties!A1:Z10000").execute()
        sheets_svc.spreadsheets().values().update(spreadsheetId=sheet_id, range="Parties!A1",
            valueInputOption="RAW", body={"values": parties_rows}).execute()
        sheets_svc.spreadsheets().values().clear(spreadsheetId=sheet_id, range="Ledger Entries!A1:Z100000").execute()
        sheets_svc.spreadsheets().values().update(spreadsheetId=sheet_id, range="Ledger Entries!A1",
            valueInputOption="RAW", body={"values": entries_rows}).execute()
        # Save last backup time and update token
        if creds.token != access_token:
            await db.users.update_one({"_id": ObjectId(current_user["_id"])}, {"$set": {"google_access_token": creds.token}})
        await db.users.update_one({"_id": ObjectId(current_user["_id"])}, {"$set": {"google_last_backup": datetime.now(timezone.utc).isoformat()}})
        sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}"
        return {"success": True, "sheet_url": sheet_url, "sheet_id": sheet_id,
                "parties_count": len(parties), "entries_count": len(entries_rows) - 1}
    except Exception as e:
        logging.error(f"Google Sheets backup error: {e}")
        raise HTTPException(500, f"Backup failed: {str(e)}")

@api_router.get("/export/sheets-status")
async def sheets_status(current_user: dict = Depends(get_current_user)):
    """Check if Google Sheets is connected for this user"""
    user = await db.users.find_one({"_id": ObjectId(current_user["_id"])})
    return {
        "connected": bool(user and user.get("google_sheets_connected")),
        "sheet_id": user.get("google_sheet_id") if user else None,
        "sheet_url": f"https://docs.google.com/spreadsheets/d/{user.get('google_sheet_id')}" if user and user.get("google_sheet_id") else None,
        "last_backup": user.get("google_last_backup") if user else None,
    }

# ─── Backup Settings + Gmail Send + APScheduler ───────────────────────────────

NEXT_BACKUP_DELTA = {"daily": timedelta(days=1), "weekly": timedelta(weeks=1), "monthly": timedelta(days=30)}

async def _get_google_creds(user: dict):
    from google.oauth2.credentials import Credentials
    return Credentials(
        token=user.get("google_access_token"),
        refresh_token=user.get("google_refresh_token"),
        token_uri="https://oauth2.googleapis.com/token",
        client_id=os.environ.get("GOOGLE_CLIENT_ID"),
        client_secret=os.environ.get("GOOGLE_CLIENT_SECRET"),
    )

async def _generate_csv_bytes(uid: str) -> bytes:
    """Generate complete CSV backup bytes for a user"""
    import csv as csv_mod
    buf = io.StringIO()
    writer = csv_mod.writer(buf)
    parties = await db.parties.find({"user_id": uid, "is_deleted": {"$ne": True}}).to_list(None)
    writer.writerow(["=== PARTIES ==="])
    writer.writerow(["ID", "Name", "Mobile", "Address", "Balance"])
    for p in parties:
        pid = str(p["_id"])
        bal = await get_party_balance(pid)
        writer.writerow([pid, p["name"], p.get("mobile",""), p.get("address",""), bal])
    writer.writerow([])
    writer.writerow(["=== LEDGER ENTRIES ==="])
    writer.writerow(["Date", "Party", "Counterparty", "Credit(Naam)", "Debit(Jama)", "Narration", "Balance", "Type", "Locked"])
    cp_cache: dict = {}
    for p in parties:
        pid = str(p["_id"])
        entries = await db.ledger_entries.find({"party_id": pid, "is_deleted": {"$ne": True}}, sort=[("date",1)]).to_list(None)
        for e in entries:
            cp_id = e.get("counterparty_id","")
            if cp_id and cp_id not in cp_cache:
                cp_doc = await db.parties.find_one({"_id": ObjectId(cp_id)})
                cp_cache[cp_id] = cp_doc["name"] if cp_doc else ""
            bal = e.get("balance", 0)
            bal_type = "Dena Hai" if bal > 0 else ("Lena Hai" if bal < 0 else "Settled")
            writer.writerow([e.get("date",""), p["name"], cp_cache.get(cp_id,""),
                             e.get("naam",0) or "", e.get("jama",0) or "",
                             e.get("narration","") or "", abs(bal), bal_type,
                             "Yes" if e.get("is_locked") else "No"])
    return buf.getvalue().encode("utf-8-sig")

async def _send_backup_via_gmail(user: dict) -> str:
    """Send backup CSV as email attachment via Gmail API. Returns recipient email."""
    from googleapiclient.discovery import build
    creds = await _get_google_creds(user)
    gmail_svc = build("gmail", "v1", credentials=creds)
    backup_email = user.get("backup_email") or user.get("email")
    if not backup_email:
        raise ValueError("No backup email configured")
    uid = str(user["_id"])
    csv_bytes = await _generate_csv_bytes(uid)
    today_str = datetime.now().strftime("%d %b %Y")
    msg = MIMEMultipart()
    msg["To"] = backup_email
    msg["From"] = "me"
    msg["Subject"] = f"PoketBook Backup — {today_str}"
    body_text = f"""Namaste!

Aapka PoketBook data backup attached hai — {today_str}

Is backup mein yeh data hai:
• Sab parties (naam, mobile, address, balance)
• Sab ledger entries (date, credit/debit, narration, balance)

Is CSV file ko Google Sheets ya Excel mein import kar sakte hain.

— Team PoketBook
poketbook.in"""
    msg.attach(MIMEText(body_text, "plain"))
    att = MIMEBase("application", "octet-stream")
    att.set_payload(csv_bytes)
    encoders.encode_base64(att)
    filename = f"poketbook_backup_{datetime.now().strftime('%Y%m%d')}.csv"
    att.add_header("Content-Disposition", f"attachment; filename={filename}")
    msg.attach(att)
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    gmail_svc.users().messages().send(userId="me", body={"raw": raw}).execute()
    # Update tokens if refreshed
    if creds.token != user.get("google_access_token"):
        await db.users.update_one({"_id": user["_id"]}, {"$set": {"google_access_token": creds.token}})
    return backup_email

@api_router.get("/backup/settings")
async def get_backup_settings(current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"_id": ObjectId(current_user["_id"])})
    return {
        "backup_email": user.get("backup_email") or user.get("email") or "",
        "backup_frequency": user.get("backup_frequency", "off"),
        "last_backup": user.get("last_email_backup"),
        "next_backup": str(user.get("next_backup_at", "")) if user.get("next_backup_at") else None,
        "google_connected": bool(user and user.get("google_sheets_connected")),
    }

@api_router.post("/backup/settings")
async def save_backup_settings(data: BackupSettings, current_user: dict = Depends(get_current_user)):
    now = datetime.now(timezone.utc)
    next_at = now + NEXT_BACKUP_DELTA.get(data.backup_frequency, timedelta(days=999))
    await db.users.update_one({"_id": ObjectId(current_user["_id"])}, {"$set": {
        "backup_email": data.backup_email,
        "backup_frequency": data.backup_frequency,
        "next_backup_at": next_at if data.backup_frequency != "off" else None,
    }})
    return {"message": "Backup settings saved", "next_backup": str(next_at) if data.backup_frequency != "off" else None}

@api_router.post("/backup/send-now")
async def send_backup_now(current_user: dict = Depends(get_current_user)):
    user = await db.users.find_one({"_id": ObjectId(current_user["_id"])})
    if not user or not user.get("google_sheets_connected"):
        raise HTTPException(400, "Google account not connected. Please connect first in the Statement page.")
    try:
        sent_to = await _send_backup_via_gmail(user)
        freq = user.get("backup_frequency", "off")
        now = datetime.now(timezone.utc)
        next_at = now + NEXT_BACKUP_DELTA.get(freq, timedelta(days=999)) if freq != "off" else None
        await db.users.update_one({"_id": ObjectId(current_user["_id"])}, {"$set": {
            "last_email_backup": now.isoformat(),
            "next_backup_at": next_at,
        }})
        return {"success": True, "sent_to": sent_to, "message": f"Backup email sent to {sent_to}"}
    except Exception as e:
        logging.error(f"Backup send error: {e}")
        raise HTTPException(500, f"Backup failed: {str(e)}")

async def _scheduled_backup_task():
    """APScheduler task — runs every hour, sends backup to due users"""
    now = datetime.now(timezone.utc)
    due_users = await db.users.find({
        "backup_frequency": {"$in": ["daily", "weekly", "monthly"]},
        "backup_email": {"$exists": True, "$ne": ""},
        "next_backup_at": {"$lte": now},
        "google_sheets_connected": True,
    }).to_list(None)
    for user in due_users:
        try:
            await _send_backup_via_gmail(user)
            freq = user.get("backup_frequency", "weekly")
            next_at = now + NEXT_BACKUP_DELTA.get(freq, timedelta(weeks=1))
            await db.users.update_one({"_id": user["_id"]}, {"$set": {
                "last_email_backup": now.isoformat(),
                "next_backup_at": next_at,
            }})
            logging.info(f"Scheduled backup sent to {user.get('backup_email')}")
        except Exception as e:
            logging.error(f"Scheduled backup failed for {user.get('_id')}: {e}")

# ─── Mount & Shutdown ─────────────────────────────────────────────────────────

app.include_router(api_router)

@app.on_event("shutdown")
async def shutdown_db_client():
    try:
        from apscheduler.schedulers.asyncio import AsyncIOScheduler
        if hasattr(app.state, "scheduler"):
            app.state.scheduler.shutdown()
    except Exception:
        pass
    client.close()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
